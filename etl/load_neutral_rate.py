"""Load the BoJ six-model natural-rate estimates (Review 2026-E-4, Chart 3) into
the tracker as the natural-rate band: low = min, mid = median, high = max across
the available models each quarter.

The committed CSV (data/seed/natural_rate_band.csv) is the source of truth so the
band reproduces without the original workbook. Re-extract from a refreshed BoJ
workbook with:  python etl/load_neutral_rate.py --xlsx path/to/Neutral_time_series.xlsx
"""
from __future__ import annotations

import argparse
import csv
import os
import statistics
from datetime import date

from db import connect, init_db, load_catalog, upsert_observations

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BAND_CSV = os.path.join(ROOT, "data", "seed", "natural_rate_band.csv")


def quarter_end(q: str) -> str:
    """'1992Q3' -> ISO date of that quarter end."""
    y = int(q[:4])
    m = {1: 3, 2: 6, 3: 9, 4: 12}[int(q[-1])]
    nd = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)
    return date.fromordinal(nd.toordinal() - 1).isoformat()


def extract_from_xlsx(path: str):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Chart 3"] if "Chart 3" in wb.sheetnames else wb.worksheets[0]
    rows = []
    for r in ws.iter_rows(values_only=True):
        q = r[2] if len(r) > 2 else None
        if not q or "Q" not in str(q):
            continue
        models = [r[i] for i in range(3, 9) if i < len(r)]
        vals = [float(v) for v in models if isinstance(v, (int, float))]
        if not vals:
            continue
        rows.append((str(q), round(min(vals), 3), round(statistics.median(vals), 3),
                     round(max(vals), 3), len(vals)))
    return rows


def write_csv(rows, path=BAND_CSV):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["quarter", "low", "mid", "high", "n_models"])
        w.writerows(rows)


def read_csv(path=BAND_CSV):
    with open(path, newline="", encoding="utf-8") as f:
        return [(r["quarter"], float(r["low"]), float(r["mid"]), float(r["high"]))
                for r in csv.DictReader(f)]


def load_band(conn, rows):
    # Clear any prior (e.g. synthetic seed) band so only the real series remains.
    conn.execute("DELETE FROM observations WHERE series_id IN "
                 "('natural_rate_low','natural_rate_mid','natural_rate_high')")
    conn.commit()
    lo = [(quarter_end(q), low) for q, low, mid, hi in rows]
    mid = [(quarter_end(q), m) for q, low, m, hi in rows]
    hi = [(quarter_end(q), h) for q, low, m, h in rows]
    n = upsert_observations(conn, "natural_rate_low", lo, "BOJ")
    n += upsert_observations(conn, "natural_rate_mid", mid, "BOJ")
    n += upsert_observations(conn, "natural_rate_high", hi, "BOJ")
    return n


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", help="re-extract band from a BoJ Chart 3 workbook")
    args = ap.parse_args()

    conn = connect()
    init_db(conn)
    load_catalog(conn)

    if args.xlsx:
        rows4 = extract_from_xlsx(args.xlsx)
        write_csv(rows4)
        print(f"Extracted {len(rows4)} quarters -> {BAND_CSV}")
        rows = [(q, lo, mi, hi) for q, lo, mi, hi, _ in rows4]
    else:
        rows = read_csv()

    n = load_band(conn, rows)
    print(f"Loaded natural-rate band: {len(rows)} quarters ({rows[0][0]}..{rows[-1][0]}), "
          f"{n} observations (low/mid/high).")


if __name__ == "__main__":
    main()
