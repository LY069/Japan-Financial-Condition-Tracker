"""Build the refreshable Excel workbook: Japan_FCI_Tracker.xlsx.

Sheets
------
Dashboard            : headline gauges, the five framework axes, auto assessment
Stage 1 - Rates      : interest-rate block (nominal curve, inflation exp, real rates)
Stage 2 - Funding    : funding costs / availability / asset prices / volumes
Data (wide)          : monthly grid, one column per series (analysis-friendly)
Indicators           : FCI / stage / axis score history + rate gap
Catalog              : full series metadata and source mapping

Refresh: re-run `python etl/run_all.py` (fetch -> indicators -> export). The
workbook is regenerated from the SQLite DB each time, so it always reflects the
latest data without manual editing.
"""
from __future__ import annotations

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from db import connect, get_series, catalog_rows, get_meta
from build_indicators import AXES, AXIS_LABEL, month_grid, to_monthly
from export_web import label_for, latest_ind, latest_obs

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "Japan_FCI_Tracker.xlsx")

NAVY = "1F3864"
BLUE = "2E5C9A"
LIGHT = "DDEBF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
AMBER = "FFEB9C"
GREY = "808080"

H1 = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
H2 = Font(name="Calibri", size=12, bold=True, color=NAVY)
BOLD = Font(bold=True)
WHITE_BOLD = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)


def score_fill(score):
    if score is None:
        return fill("F2F2F2")
    if score >= 0.15:
        return fill(GREEN)
    if score > -0.15:
        return fill(AMBER)
    return fill(RED)


def style_header_row(ws, row, ncols, color=BLUE):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(color)
        cell.font = WHITE_BOLD
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def title_banner(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.fill = fill(NAVY)
    c.font = H1
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30


def build_dashboard(wb, conn):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    title_banner(ws, "Japan Financial Conditions Tracker  —  BoJ two-stage framework", 6)

    fci = latest_ind(conn, "composite", "fci")
    rg = latest_ind(conn, "composite", "rate_gap")
    latest = fci["date"] if fci else ""

    def lv(sid):
        _, v = latest_obs(conn, sid)
        return v

    ws["A3"] = "As of"; ws["A3"].font = BOLD
    ws["B3"] = latest
    ws["D3"] = "Data mode"; ws["D3"].font = BOLD
    ws["E3"] = get_meta(conn, "data_mode", "")

    # Headline metric cards
    cards = [
        ("Policy rate", f"{lv('policy_rate'):.2f}%"),
        ("Composite FCI (z)", f"{fci['score']:+.2f}" if fci else "n/a"),
        ("FCI assessment", label_for(fci["score"]) if fci else "n/a"),
        ("Real policy rate", f"{rg['value']:+.2f}%" if rg else "n/a"),
        ("Natural rate (mid)", f"{lv('natural_rate_mid'):+.2f}%"),
        ("Rate gap (real - r*)", f"{rg['score']:+.2f}pp" if rg else "n/a"),
    ]
    r = 5
    for i, (lab, val) in enumerate(cards):
        col = 1 + (i % 3) * 2
        row = r + (i // 3) * 3
        ws.cell(row=row, column=col, value=lab).font = Font(bold=True, color=GREY, size=9)
        vc = ws.cell(row=row + 1, column=col, value=val)
        vc.font = Font(bold=True, size=14, color=NAVY)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)

    # Axis table
    ar = 12
    ws.cell(row=ar, column=1, value="Framework axes").font = H2
    ar += 1
    ws.append  # noop
    headers = ["Axis", "Accommodation score (z)", "Assessment"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=ar, column=j, value=h)
    style_header_row(ws, ar, 3)
    for a in ["real_rate", "funding_costs", "availability", "asset_prices", "funding_volumes"]:
        li = latest_ind(conn, "axis", a)
        sc = li["score"] if li else None
        ar += 1
        ws.cell(row=ar, column=1, value=AXIS_LABEL[a]).border = BORDER
        c = ws.cell(row=ar, column=2, value=round(sc, 2) if sc is not None else None)
        c.border = BORDER; c.fill = score_fill(sc); c.alignment = Alignment(horizontal="center")
        c3 = ws.cell(row=ar, column=3, value=label_for(sc))
        c3.border = BORDER; c3.fill = score_fill(sc)

    # Stage composites
    ar += 2
    ws.cell(row=ar, column=1, value="Stage composites").font = H2
    ar += 1
    for st, lab in [("stage1", "Stage 1 - Interest rate developments"),
                    ("stage2", "Stage 2 - Funding environment")]:
        li = latest_ind(conn, "stage", st)
        sc = li["score"] if li else None
        ws.cell(row=ar, column=1, value=lab).border = BORDER
        c = ws.cell(row=ar, column=2, value=round(sc, 2) if sc is not None else None)
        c.border = BORDER; c.fill = score_fill(sc); c.alignment = Alignment(horizontal="center")
        ws.cell(row=ar, column=3, value=label_for(sc)).border = BORDER
        ar += 1

    # Assessment text
    ar += 1
    ws.cell(row=ar, column=1, value="Assessment").font = H2
    ar += 1
    from export_web import build_assessment
    headline = {
        "policy_rate": lv("policy_rate"), "real_policy_rate": rg["value"] if rg else None,
        "natural_rate_mid": lv("natural_rate_mid"), "rate_gap": rg["score"] if rg else None,
        "fci_score": fci["score"] if fci else None,
        "natural_rate_low": lv("natural_rate_low"), "natural_rate_high": lv("natural_rate_high"),
    }
    txt = build_assessment(conn, headline)
    ws.merge_cells(start_row=ar, start_column=1, end_row=ar + 4, end_column=6)
    tc = ws.cell(row=ar, column=1, value=txt)
    tc.alignment = Alignment(wrap_text=True, vertical="top")
    tc.fill = fill(LIGHT)

    # FCI history mini-chart
    chart_anchor = ar + 6
    rows = conn.execute("SELECT date,score FROM indicators WHERE scope='composite' AND key='fci' "
                        "ORDER BY date").fetchall()
    # write helper data far to the right
    hc = 9
    ws.cell(row=1, column=hc, value="date")
    ws.cell(row=1, column=hc + 1, value="FCI")
    for i, rr in enumerate(rows, start=2):
        ws.cell(row=i, column=hc, value=rr["date"])
        ws.cell(row=i, column=hc + 1, value=round(rr["score"], 3))
    chart = LineChart()
    chart.title = "Composite Financial Conditions Index (z; + = accommodative)"
    chart.height = 7; chart.width = 18
    data = Reference(ws, min_col=hc + 1, min_row=1, max_row=len(rows) + 1)
    cats = Reference(ws, min_col=hc, min_row=2, max_row=len(rows) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, f"A{chart_anchor}")

    widths = {"A": 34, "B": 22, "C": 22, "D": 16, "E": 18, "F": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions[get_column_letter(hc)].hidden = True
    ws.column_dimensions[get_column_letter(hc + 1)].hidden = True


def build_block_sheet(wb, conn, title, sids):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    title_banner(ws, title, 6)
    headers = ["Series", "Latest date", "Latest value", "Unit", "Accommodation (z)", "Assessment"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=3, column=j, value=h)
    style_header_row(ws, 3, len(headers))
    cat = {r["series_id"]: r for r in catalog_rows(conn)}
    r = 4
    for sid in sids:
        if sid not in cat:
            continue
        meta = cat[sid]
        d, v = latest_obs(conn, sid)
        li = latest_ind(conn, "series", sid)
        sc = li["score"] if li else None
        ws.cell(row=r, column=1, value=meta["name"]).border = BORDER
        ws.cell(row=r, column=2, value=d).border = BORDER
        cv = ws.cell(row=r, column=3, value=round(v, 3) if v is not None else None); cv.border = BORDER
        ws.cell(row=r, column=4, value=meta["unit"]).border = BORDER
        cs = ws.cell(row=r, column=5, value=round(sc, 2) if sc is not None else None)
        cs.border = BORDER; cs.fill = score_fill(sc); cs.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6, value=label_for(sc) if sc is not None else "context").border = BORDER
        r += 1
    for col, w in zip("ABCDEF", [42, 12, 14, 14, 16, 20]):
        ws.column_dimensions[col].width = w


def build_data_wide(wb, conn):
    ws = wb.create_sheet("Data (wide)")
    grid = month_grid(end=None)
    cat = list(catalog_rows(conn))
    sids = [r["series_id"] for r in cat]
    ws.cell(row=1, column=1, value="date")
    for j, sid in enumerate(sids, start=2):
        ws.cell(row=1, column=j, value=sid)
    style_header_row(ws, 1, len(sids) + 1, color=NAVY)
    monthly = {sid: to_monthly(get_series(conn, sid), grid) for sid in sids}
    for i, d in enumerate(grid, start=2):
        ws.cell(row=i, column=1, value=d)
        for j, sid in enumerate(sids, start=2):
            v = monthly[sid].get(d)
            if v is not None:
                ws.cell(row=i, column=j, value=round(v, 4))
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 12
    for j in range(2, len(sids) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 14


def build_indicators_sheet(wb, conn):
    ws = wb.create_sheet("Indicators")
    grid = month_grid(end=None)
    keys = [("composite", "fci", "FCI"), ("composite", "rate_gap", "rate_gap"),
            ("stage", "stage1", "stage1"), ("stage", "stage2", "stage2")]
    keys += [("axis", a, a) for a in AXES]
    ws.cell(row=1, column=1, value="date")
    for j, (_, _, lab) in enumerate(keys, start=2):
        ws.cell(row=1, column=j, value=lab)
    style_header_row(ws, 1, len(keys) + 1, color=NAVY)
    data = {}
    for scope, key, _ in keys:
        rows = conn.execute("SELECT date,score FROM indicators WHERE scope=? AND key=? ORDER BY date",
                            (scope, key)).fetchall()
        data[(scope, key)] = {r["date"]: r["score"] for r in rows}
    for i, d in enumerate(grid, start=2):
        ws.cell(row=i, column=1, value=d)
        for j, (scope, key, _) in enumerate(keys, start=2):
            v = data[(scope, key)].get(d)
            if v is not None:
                ws.cell(row=i, column=j, value=round(v, 4))
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 12
    for j in range(2, len(keys) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 13


def build_catalog_sheet(wb, conn):
    ws = wb.create_sheet("Catalog")
    cols = ["series_id", "name", "stage", "category", "subcategory", "unit", "frequency",
            "polarity", "weight", "source", "source_series_id", "source_url", "notes"]
    for j, c in enumerate(cols, start=1):
        ws.cell(row=1, column=j, value=c)
    style_header_row(ws, 1, len(cols), color=NAVY)
    for i, r in enumerate(catalog_rows(conn), start=2):
        for j, c in enumerate(cols, start=1):
            ws.cell(row=i, column=j, value=r[c])
    ws.freeze_panes = "A2"
    for col, w in zip("ABCDEFGHIJKLM", [18, 40, 20, 16, 16, 12, 10, 9, 8, 10, 18, 40, 50]):
        ws.column_dimensions[col].width = w


def main():
    conn = connect()
    wb = Workbook()
    build_dashboard(wb, conn)
    build_block_sheet(wb, conn, "Stage 1 - Rates",
                      ["policy_rate", "jgb_1y", "jgb_2y", "jgb_5y", "jgb_10y", "jgb_30y",
                       "infexp_1y", "infexp_3y", "infexp_10y", "real_1y", "real_3y", "real_10y",
                       "real_policy_rate"])
    build_block_sheet(wb, conn, "Stage 2 - Funding",
                      ["lending_rate", "cp_rate", "corp_bond_spread",
                       "tankan_lend_large", "tankan_lend_small", "tankan_finpos_large",
                       "tankan_finpos_small", "topix", "nikkei225", "usdjpy",
                       "bank_lending_yoy", "cp_corpbond_yoy"])
    build_data_wide(wb, conn)
    build_indicators_sheet(wb, conn)
    build_catalog_sheet(wb, conn)
    wb.save(OUT)
    print(f"Wrote {OUT} ({os.path.getsize(OUT)//1024} KB).")


if __name__ == "__main__":
    main()
